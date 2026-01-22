"""
Processeur pour le suivi des stocks
Adapté depuis votre script original
"""

import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
import tempfile
import os


def process(files: dict, params: dict) -> str:
    """
    Traite les fichiers de suivi des stocks
    """
    print("=" * 50)
    print("🚀 DÉBUT DU TRAITEMENT")
    print("=" * 50)
    
    file_tracking = files['tracking']
    file_export = files['export']
    export_date_str = params['export_date']
    
    print(f"📁 Fichier tracking: {file_tracking}")
    print(f"📁 Fichier export: {file_export}")
    print(f"📅 Date: {export_date_str}")
    
    # Accepter les deux formats de date
    try:
        export_date = datetime.datetime.strptime(export_date_str, '%d/%m/%Y')
        print(f"✅ Date parsée (format FR): {export_date}")
    except ValueError:
        try:
            export_date = datetime.datetime.strptime(export_date_str, '%Y-%m-%d')
            print(f"✅ Date parsée (format ISO): {export_date}")
        except ValueError:
            raise ValueError("Format de date invalide. Utilisez jj/mm/aaaa ou YYYY-MM-DD")
    
    def progress_callback(current, total):
        pass
    
    # Exécuter les mises à jour avec logs
    try:
        print("📊 Étape 1/3: update_tracking...")
        update_tracking(file_tracking, file_export, export_date, progress_callback)
        print("✅ update_tracking terminé")
        
        print("📊 Étape 2/3: update_monthly_tracking...")
        update_monthly_tracking(file_tracking, export_date, progress_callback)
        print("✅ update_monthly_tracking terminé")
        
        print("📊 Étape 3/3: update_semestrial_tracking...")
        update_semestrial_tracking(file_tracking, export_date, progress_callback)
        print("✅ update_semestrial_tracking terminé")
        
    except Exception as e:
        print(f"❌ ERREUR: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    print("=" * 50)
    print(f"🎉 TRAITEMENT TERMINÉ - Fichier: {file_tracking}")
    print("=" * 50)
    
    return file_tracking


def find_sheet(workbook, prefix):
    """Trouve une feuille par préfixe"""
    prefix = prefix.lower()
    for sheet in workbook.sheetnames:
        if sheet.lower().startswith(prefix):
            return sheet
    return None


def clear_worksheet(worksheet):
    """Nettoie une feuille de calcul"""
    for merge in list(worksheet.merged_cells):
        worksheet.unmerge_cells(str(merge))
    max_row = worksheet.max_row
    if max_row > 1:
        worksheet.delete_rows(2, max_row - 1)


def style_headers(worksheet):
    """Style les en-têtes"""
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    adjust_column_width(worksheet)


def adjust_column_width(worksheet):
    """Ajuste la largeur des colonnes"""
    for column_cells in worksheet.columns:
        length = 0
        for cell in column_cells:
            try:
                value_length = len(str(cell.value))
            except:
                value_length = 0
            if value_length > length:
                length = value_length
        
        adjusted_length = min(length + 2, 50)
        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_length


def format_date_columns(df):
    """Formate les colonnes de dates"""
    date_columns = df.select_dtypes(include=['datetime64[ns]', 'datetime']).columns
    for col in date_columns:
        df[col] = df[col].dt.strftime('%d/%m/%Y')
    return df


def is_valid_date(date_str):
    """Vérifie si une date est valide"""
    try:
        datetime.datetime.strptime(date_str, '%d/%m/%Y')
        return True
    except ValueError:
        return False


def update_tracking(file_tracking, file_export, export_date, progress_callback):
    """Met à jour le suivi des stocks"""
    df_tracking = pd.read_excel(file_tracking, sheet_name='Liste de Stock')
    df_export = pd.read_excel(file_export)
    
    export_date_str = export_date.strftime('%d/%m/%Y')
    
    if export_date_str in df_tracking.columns:
        raise ValueError("Les données pour cette date ont déjà été importées.")
    
    # Regrouper par Code article + Emplacement
    df_grouped = df_export.groupby(['Code article', 'Emplacement']).size().reset_index(name='Quantité')
    
    df_descriptions = df_export.groupby(['Code article', 'Emplacement']).agg({
        "Description de l'actif": 'first',
        "Description de l'emplacement": 'first'
    }).reset_index()
    
    df_grouped = df_grouped.merge(df_descriptions, on=['Code article', 'Emplacement'], how='left')
    df_tracking[export_date_str] = 0
    
    total_operations = len(df_grouped) + len(df_tracking)
    processed_operations = 0
    
    for idx, row in df_grouped.iterrows():
        codification = row['Code article']
        magasin = row['Emplacement']
        quantite = row['Quantité']
        designation = row["Description de l'actif"]
        description_emplacement = row["Description de l'emplacement"]
        
        condition = (df_tracking['Codification DSNA'] == codification) & (df_tracking['Magasin'] == magasin)
        if not df_tracking[condition].empty:
            df_tracking.loc[condition, export_date_str] = quantite
        else:
            new_row = {
                'Codification DSNA': codification,
                'Désignation': designation,
                'Magasin': magasin,
                'Description': description_emplacement,
                export_date_str: quantite
            }
            new_row_df = pd.DataFrame([new_row])
            df_tracking = pd.concat([df_tracking, new_row_df], ignore_index=True)
        
        processed_operations += 1
        progress_callback(processed_operations, total_operations)
    
    df_tracking = format_date_columns(df_tracking)
    
    # CORRECTION: Tout faire en une seule opération avec le workbook
    workbook = load_workbook(file_tracking)
    
    # Supprimer et recréer la feuille
    if 'Liste de Stock' in workbook.sheetnames:
        del workbook['Liste de Stock']
    
    worksheet = workbook.create_sheet('Liste de Stock', 0)
    
    # Écrire les données
    for r_idx, row in enumerate(dataframe_to_rows(df_tracking, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Appliquer le style
    style_headers(worksheet)
    
    # Sauvegarder et fermer
    workbook.save(file_tracking)
    workbook.close()


def update_monthly_tracking(file_tracking, export_date, progress_callback):
    """Met à jour le suivi mensuel"""
    # CORRECTION: Charger le workbook une seule fois
    workbook = load_workbook(file_tracking)
    
    try:
        suivi_mensuel_sheet_name = find_sheet(workbook, 'suivi mensuel')
        
        if not suivi_mensuel_sheet_name:
            workbook.close()
            raise ValueError("Aucun onglet 'Suivi Mensuel' trouvé")
        
        worksheet = workbook[suivi_mensuel_sheet_name]
        clear_worksheet(worksheet)
        
        df_tracking = pd.read_excel(file_tracking, sheet_name='Liste de Stock')
        export_date_str = export_date.strftime('%d/%m/%Y')
        
        if export_date_str not in df_tracking.columns:
            worksheet.insert_rows(2)
            worksheet['A3'] = "Pas de données disponibles pour le mois."
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            workbook.save(file_tracking)
            return
        
        col_idx = df_tracking.columns.get_loc(export_date_str)
        
        if col_idx < 1:
            worksheet.insert_rows(2)
            worksheet['A3'] = "Pas de données disponibles pour le mois."
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            workbook.save(file_tracking)
            return
        
        previous_month_col = df_tracking.columns[col_idx - 1]
        
        if not is_valid_date(previous_month_col):
            worksheet.insert_rows(2)
            worksheet['A3'] = "Pas de données disponibles pour le mois."
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            workbook.save(file_tracking)
            return
        
        df_tracking['Variation'] = df_tracking[export_date_str] - df_tracking[previous_month_col]
        df_tracking['Quantité actuelle'] = df_tracking[export_date_str]
        
        variations = df_tracking[df_tracking['Variation'] != 0]
        variations = variations[['Codification DSNA', 'Désignation', 'Magasin', 'Description', 'Variation', 'Quantité actuelle']]
        
        start_date = datetime.datetime.strptime(previous_month_col, '%d/%m/%Y').strftime('%d/%m/%Y')
        end_date = datetime.datetime.strptime(export_date_str, '%d/%m/%Y').strftime('%d/%m/%Y')
        period_study = f"Variation entre le {start_date} et le {end_date}"
        
        worksheet.insert_rows(2)
        worksheet['A1'] = period_study
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        if not variations.empty:
            for r_idx, row in enumerate(dataframe_to_rows(variations, index=False, header=False), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                    if pd.notna(value) and c_idx == len(variations.columns):  # Dernière colonne
                        if value <= 5:
                            cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                        elif value <= 10:
                            cell.fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
        else:
            worksheet.insert_rows(3)
            worksheet['A3'] = "Aucune variation ce mois-ci"
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        headers = ['Codification DSNA', 'Désignation', 'Magasin', 'Description', 'Variation', 'Quantité actuelle']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col_num, value=header)
            cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(border_style='thin', color='FFFFFF'),
                right=Side(border_style='thin', color='FFFFFF'),
                top=Side(border_style='thin', color='FFFFFF'),
                bottom=Side(border_style='thin', color='FFFFFF')
            )
        
        # CORRECTION: Sauvegarder et fermer
        workbook.save(file_tracking)
        
    finally:
        # IMPORTANT: Toujours fermer le workbook
        workbook.close()


def update_semestrial_tracking(file_tracking, export_date, progress_callback):
    """Met à jour le suivi semestriel"""
    # CORRECTION: Même pattern que monthly
    workbook = load_workbook(file_tracking)
    
    try:
        suivi_semestrial_sheet_name = find_sheet(workbook, 'suivi semestriel')
        
        if not suivi_semestrial_sheet_name:
            workbook.close()
            raise ValueError("Aucun onglet 'Suivi Semestriel' trouvé")
        
        worksheet = workbook[suivi_semestrial_sheet_name]
        clear_worksheet(worksheet)
        
        df_tracking = pd.read_excel(file_tracking, sheet_name='Liste de Stock')
        export_date_str = export_date.strftime('%d/%m/%Y')
        
        if export_date_str not in df_tracking.columns:
            worksheet.insert_rows(2)
            worksheet['A3'] = "Pas de données disponibles pour le semestre."
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            workbook.save(file_tracking)
            return
        
        col_idx = df_tracking.columns.get_loc(export_date_str)
        
        if col_idx < 6:
            worksheet.insert_rows(2)
            worksheet['A3'] = "Pas de données disponibles pour le semestre."
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            workbook.save(file_tracking)
            return
        
        previous_semester_col = df_tracking.columns[col_idx - 6]
        
        if not is_valid_date(previous_semester_col):
            worksheet.insert_rows(2)
            worksheet['A3'] = "Pas de données disponibles pour le semestre."
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            workbook.save(file_tracking)
            return
        
        df_tracking['Variation'] = df_tracking[export_date_str] - df_tracking[previous_semester_col]
        df_tracking['Quantité actuelle'] = df_tracking[export_date_str]
        
        variations = df_tracking[df_tracking['Variation'] != 0]
        variations = variations[['Codification DSNA', 'Désignation', 'Magasin', 'Description', 'Variation', 'Quantité actuelle']]
        
        start_date = datetime.datetime.strptime(previous_semester_col, '%d/%m/%Y').strftime('%d/%m/%Y')
        end_date = datetime.datetime.strptime(export_date_str, '%d/%m/%Y').strftime('%d/%m/%Y')
        period_study = f"Variation entre le {start_date} et le {end_date}"
        
        worksheet.insert_rows(2)
        worksheet['A1'] = period_study
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        if not variations.empty:
            for r_idx, row in enumerate(dataframe_to_rows(variations, index=False, header=False), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                    if pd.notna(value) and c_idx == len(variations.columns):
                        if value <= 5:
                            cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                        elif value <= 10:
                            cell.fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
        else:
            worksheet.insert_rows(3)
            worksheet['A3'] = "Aucune variation ce semestre"
            worksheet.merge_cells('A3:F3')
            worksheet['A3'].alignment = Alignment(horizontal='center')
            worksheet['A3'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        headers = ['Codification DSNA', 'Désignation', 'Magasin', 'Description', 'Variation', 'Quantité actuelle']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col_num, value=header)
            cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(border_style='thin', color='FFFFFF'),
                right=Side(border_style='thin', color='FFFFFF'),
                top=Side(border_style='thin', color='FFFFFF'),
                bottom=Side(border_style='thin', color='FFFFFF')
            )
        
        # CORRECTION: Sauvegarder et fermer
        workbook.save(file_tracking)
        
    finally:
        # IMPORTANT: Toujours fermer le workbook
        workbook.close()
